from __future__ import annotations

import math
import re
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .config import DATA_DIR
from .knowledge import KnowledgeBase, normalize, normalize_key
from .models import QuoteItem


def local_price_file() -> Path:
    files = sorted(DATA_DIR.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError("В папке «Локальные данные» нет XLSX-прайса")
    return files[0]


def _fabric_catalog(price_path: Path, db: KnowledgeBase) -> list[dict[str, str]]:
    sheet = load_workbook(price_path, data_only=True, read_only=True)["Рулонные ткани"]
    result: list[dict[str, str]] = []
    for row in range(5, sheet.max_row + 1):
        collection = normalize(sheet.cell(row, 3).value)
        category = normalize(sheet.cell(row, 5).value).upper()
        if collection and category in {"Е", "E", "1", "2", "3", "4", "5"}:
            category = "E" if category in {"Е", "E"} else category
            result.append({"collection": collection, "category": category})
            db.put("fabric", collection, {"collection": collection, "category": category}, "Локальный прайс / Рулонные ткани", 1.0, True)
    return result


def _category(item: QuoteItem, catalog: list[dict[str, str]], db: KnowledgeBase) -> tuple[str | None, str | None, str]:
    query = normalize_key(item.fabric)
    candidates = [row for row in catalog if normalize_key(row["collection"]) in query or query in normalize_key(row["collection"])] if query else []
    if len(candidates) == 1:
        return candidates[0]["category"], None, f"Локальный прайс: ткань {candidates[0]['collection']}"
    if query.startswith("АЛЬФА"):
        return ("2" if "БЛЭКАУТ" in normalize_key(item.opacity) else "E"), None, "Локальный прайс: правило коллекции Альфа"
    if "СКРИН 5%" in query:
        return "3", None, "Локальный прайс: коллекция СКРИН 5%"
    if "ОМЕГА" in query and ("BLACK-OUT" in query or "BLACKOUT" in query or "БЛЭКАУТ" in query):
        return "2", None, "Локальный прайс: коллекция ОМЕГА BLACK-OUT"
    analogue = db.search("analogue", item.fabric, limit=1, verified_only=True)
    if analogue:
        data = analogue[0]["payload"]
        if data.get("category") and data.get("local_fabric"):
            return str(data["category"]), str(data["local_fabric"]), f"Подтверждённый аналог: {data.get('source_url', '')}"
    return None, None, ""


def _system_sheet(system: str) -> str | None:
    value = normalize_key(system)
    if "СТАНДАРТ" in value:
        return "Стандарт MG"
    if "МИНИ" in value or "MINI" in value:
        return "Mini"
    if "UNI 1" in value or "UNI-1" in value:
        return "Uni 1 "
    if "UNI 2" in value or "UNI-2" in value:
        return "Uni 2"
    if "AMG" in value:
        return "AMG"
    return None


def _ceil(value: float, grid: list[float]) -> float | None:
    return next((point for point in grid if point + 1e-9 >= value), None)


def matrix_price(price_path: Path, sheet_name: str, category: str, height: float, width: float) -> tuple[float | None, str]:
    return _matrix_price_cached(str(price_path), price_path.stat().st_mtime_ns, sheet_name, category, height, width)


@lru_cache(maxsize=1024)
def _matrix_price_cached(price_path_str: str, _price_version: int, sheet_name: str, category: str, height: float, width: float) -> tuple[float | None, str]:
    price_path = Path(price_path_str)
    sheet = load_workbook(price_path, data_only=True, read_only=True)[sheet_name]
    labels = {"E": ("е категория", "e категория"), "1": ("1 категория",), "2": ("2 категория",), "3": ("3 категория",), "4": ("4 категория",), "5": ("5 категория",)}
    marker_row = None
    for row in range(1, sheet.max_row + 1):
        text = normalize(sheet.cell(row, 2).value).lower()
        if any(label in text for label in labels[category]):
            marker_row = row
            break
    if marker_row is None:
        return None, ""
    header_row = marker_row + 1
    widths: list[tuple[int, float]] = []
    for col in range(3, sheet.max_column + 1):
        value = sheet.cell(header_row, col).value
        if isinstance(value, (int, float)):
            widths.append((col, float(value)))
        elif widths:
            break
    heights: list[tuple[int, float]] = []
    for row in range(marker_row + 2, sheet.max_row + 1):
        value = sheet.cell(row, 2).value
        if isinstance(value, (int, float)):
            heights.append((row, float(value)))
        elif heights:
            break
    selected_width = _ceil(width, [value for _, value in widths])
    selected_height = _ceil(height, [value for _, value in heights])
    if selected_width is None or selected_height is None:
        return None, ""
    col = next(index for index, value in widths if value == selected_width)
    row = next(index for index, value in heights if value == selected_height)
    price = sheet.cell(row, col).value
    if not isinstance(price, (int, float)):
        return None, ""
    provenance = f"{price_path.name} / {sheet_name} / категория {category} / сетка {selected_width:.2f}x{selected_height:.2f} м"
    return float(price), provenance


def vertical_price(price_path: Path, item: QuoteItem, usd_rub_rate: float) -> tuple[int | None, str | None, str]:
    query = normalize_key(f"{item.name} {item.fabric}")
    if "ЖАЛЮЗИ" not in query or item.area_m2 is None:
        return None, None, ""
    if "СТЕНОВ" in query and "КРОНШТЕЙН" in query:
        return 0, "комплектация", "Подтверждённый прецедент: ПримерыКП / КП АО «Трансинжстрой» / стеновые кронштейны без доплаты"
    sheet = load_workbook(price_path, data_only=True, read_only=True)["Вертикальные"]
    for row in range(7, 13):
        collections = normalize_key(sheet.cell(row, 2).value)
        rate = sheet.cell(row, 6).value
        category = normalize(sheet.cell(row, 7).value).upper()
        if not isinstance(rate, (int, float)):
            continue
        names = []
        for raw_name in re.split(r"[,;]", collections):
            name = re.sub(r"\bII\b", "", normalize_key(raw_name)).strip()
            if len(name) >= 4:
                names.append(name)
        if not any(name in query or query in name for name in names):
            continue
        billable_area = max(1.0, float(item.area_m2))
        price = round(billable_area * float(rate) * usd_rub_rate)
        source = f"{price_path.name} / Вертикальные / строка {row} / категория {category} / {rate:.4f} $/м² / курс {usd_rub_rate} руб."
        return price, category, source
    return None, None, ""


def _usd_option(price_path: Path, sheet_name: str, cell: str) -> float:
    return _usd_option_cached(str(price_path), price_path.stat().st_mtime_ns, sheet_name, cell)


@lru_cache(maxsize=256)
def _usd_option_cached(price_path_str: str, _price_version: int, sheet_name: str, cell: str) -> float:
    value = load_workbook(Path(price_path_str), data_only=True, read_only=True)[sheet_name][cell].value
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"\d+(?:[.,]\d+)?", normalize(value))
    if not match:
        raise ValueError(f"Не удалось прочитать цену опции {sheet_name}!{cell}")
    return float(match.group().replace(",", "."))


def bnt_price(price_path: Path, item: QuoteItem, category: str, usd_rub_rate: float) -> tuple[int | None, str]:
    variant = item.raw.get("variant")
    if variant == "bnt_m44_mono_electric":
        widths = [float(value) for value in item.raw.get("component_widths_m", [])]
        if not widths or item.height_m is None:
            return None, ""
        bases: list[float] = []
        sources: list[str] = []
        for width in widths:
            base, source = matrix_price(price_path, "BEN M", category, float(item.height_m), width)
            if base is None:
                return None, ""
            bases.append(base)
            sources.append(source)
        total_width = sum(widths)
        mono = _usd_option(price_path, "BEN M", "H31") * len(widths)
        tube_44 = _usd_option(price_path, "BEN M", "H25") * total_width
        mounting_profile = _usd_option(price_path, "BEN M", "H37") * total_width
        electric = _usd_option(price_path, "Электрика AMIGO", "E35")
        total_usd = sum(bases) + mono + tube_44 + mounting_profile + electric
        provenance = (
            "; ".join(sources)
            + f" + BEN M!H31 MONO {mono:.2f} $ + BEN M!H25 труба 44 {tube_44:.2f} $"
            + f" + BEN M!H37 монтажный профиль {mounting_profile:.2f} $"
            + f" + Электрика AMIGO!E35 {electric:.2f} $ / курс {usd_rub_rate} руб."
        )
        return round(total_usd * usd_rub_rate), provenance
    if variant == "bnt_l65_electric":
        width = float(item.width_m or 0)
        if not width or item.height_m is None:
            return None, ""
        base, source = matrix_price(price_path, "BEN L", category, float(item.height_m), width)
        if base is None:
            return None, ""
        tube_65 = _usd_option(price_path, "BEN L", "H27") * width
        mounting_profile = _usd_option(price_path, "BEN L", "H37") * width
        electric = _usd_option(price_path, "Электрика AMIGO", "E37")
        total_usd = base + tube_65 + mounting_profile + electric
        provenance = (
            f"{source} + BEN L!H27 труба 65 {tube_65:.2f} $"
            + f" + BEN L!H37 монтажный профиль {mounting_profile:.2f} $"
            + f" + Электрика AMIGO!E37 {electric:.2f} $ / курс {usd_rub_rate} руб."
        )
        return round(total_usd * usd_rub_rate), provenance
    return None, ""


def accessory_price(price_path: Path, item: QuoteItem, usd_rub_rate: float) -> tuple[int | None, str]:
    cells = {"amigo_remote_1ch": "E84", "amigo_remote_5ch": "E85"}
    cell = cells.get(str(item.raw.get("variant")))
    if not cell:
        return None, ""
    value = _usd_option(price_path, "Электрика AMIGO", cell)
    return round(value * usd_rub_rate), f"{price_path.name} / Электрика AMIGO!{cell} / {value:.4f} $ / курс {usd_rub_rate} руб."


def price_items(items: list[QuoteItem], config: dict[str, Any], db: KnowledgeBase, logger=print) -> tuple[list[QuoteItem], list[QuoteItem], list[QuoteItem]]:
    price_path = local_price_file()
    catalog = _fabric_catalog(price_path, db)
    priced: list[QuoteItem] = []
    unresolved: list[QuoteItem] = []
    invalid: list[QuoteItem] = []
    for item in items:
        if not item.has_size:
            item.note = "Нет размеров или площади"
            invalid.append(item)
            continue
        if item.raw.get("variant") == "angular_unverified":
            item.note = "В локальном прайсе не найдено подтверждённое правило расчёта угловой шторы"
            unresolved.append(item)
            continue
        accessory, accessory_source = accessory_price(price_path, item, float(config["usd_rub_rate"]))
        if accessory is not None:
            item.price_rub = accessory
            item.category = "комплектация"
            item.price_source = accessory_source
            priced.append(item)
            logger(f"{item.source_ref}: электрика, {accessory} руб./ед.")
            continue
        if item.area_m2 and not (item.width_m and item.height_m):
            price, category, provenance = vertical_price(price_path, item, float(config["usd_rub_rate"]))
            if price is None:
                item.note = "Площадь распознана, но для этой позиции нет проверенного правила цены за м²"
                unresolved.append(item)
                continue
            item.price_rub = price
            item.category = category
            item.price_source = provenance
            priced.append(item)
            logger(f"{item.source_ref}: вертикальные жалюзи, категория {category}, {price} руб./ед.")
            continue
        if item.raw.get("variant") in {"bnt_m44_mono_electric", "bnt_l65_electric"}:
            category, _, fabric_source = _category(item, catalog, db)
            if not category:
                item.note = f"Не найдена ценовая категория ткани «{item.fabric}»"
                unresolved.append(item)
                continue
            item.category = category
            price, provenance = bnt_price(price_path, item, category, float(config["usd_rub_rate"]))
            if price is None:
                item.note = "Размер вне проверенной сетки BNT или отсутствует правило комплектации"
                unresolved.append(item)
                continue
            item.price_rub = price
            item.price_source = provenance
            priced.append(item)
            logger(f"{item.source_ref}: {item.system}, категория {category}, электрика, {price} руб./ед.")
            continue
        sheet_name = _system_sheet(item.system)
        if not sheet_name:
            item.note = "Не определена поддерживаемая система изделия"
            unresolved.append(item)
            continue
        category, local_fabric, fabric_source = _category(item, catalog, db)
        if not category:
            item.note = f"Не найдена ценовая категория ткани «{item.fabric}»"
            unresolved.append(item)
            continue
        item.category = category
        if local_fabric:
            original = item.fabric
            item.fabric = local_fabric
            item.analogue_source = fabric_source
            item.note = f"Аналог: {original} → {local_fabric}"
        split = max(1, int(item.split_into or 1))
        width = float(item.width_m or 0)
        if sheet_name == "Стандарт MG":
            width -= 0.04
        base, provenance = matrix_price(price_path, sheet_name, category, float(item.height_m), width / split)
        if base is None and split == 1 and sheet_name in {"Mini", "Стандарт MG"}:
            replacements = []
            if sheet_name == "Mini":
                replacements.append(("Стандарт MG", float(item.width_m) - 0.04, "Стандарт"))
            replacements.append(("AMG", float(item.width_m), "AMG"))
            for replacement_sheet, replacement_width, display_system in replacements:
                base, provenance = matrix_price(price_path, replacement_sheet, category, float(item.height_m), replacement_width)
                if base is not None:
                    item.note = (item.note + "; " if item.note else "") + f"Система по габариту: {item.system} → {display_system}"
                    item.system = display_system
                    break
        if base is None:
            item.note = "Размер вне проверенной ценовой сетки"
            unresolved.append(item)
            continue
        if sheet_name == "AMG" and item.raw.get("variant") == "cassette_32_guides":
            cassette = 45.66 * float(item.width_m)
            side_guides = 56.45 * float(item.height_m)
            base += cassette + side_guides
            provenance += f" + кассета 32 мм {cassette:.2f} $ + боковые направляющие {side_guides:.2f} $"
        item.price_rub = round(base * float(config["usd_rub_rate"])) * split
        item.price_source = provenance + f" / курс {config['usd_rub_rate']} руб."
        priced.append(item)
        logger(f"{item.source_ref}: {item.system}, категория {category}, {item.price_rub} руб./ед.")
    return priced, unresolved, invalid
